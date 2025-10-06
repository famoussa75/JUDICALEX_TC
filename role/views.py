import html
from django.http import HttpResponse,JsonResponse
from django.views.decorators.csrf import csrf_exempt
import json
from django.contrib import messages
from django.core.paginator import Paginator
from django.shortcuts import render,redirect,get_object_or_404
from _base.models import Juridictions
from magistrats.models import Presidents
from .forms import RoleForm,RoleAffaireForm,EnrollementForm,DecisionsForm,MessageForm
from django.db import IntegrityError, transaction
from django.forms import inlineformset_factory, modelformset_factory
from .models import AffaireRoles, DecisionHistory, EnrollementHistory, Roles, Enrollement, Decisions, SuivreAffaire
from datetime import datetime, timedelta, date
from django.db.models import Count, Case, When, Value, CharField, Q, F, OuterRef, Subquery
from django.utils.html import mark_safe
import re
from time import sleep
from account.models import Account, Notification
import openpyxl

from reportlab.lib.pagesizes import A4, landscape
from reportlab.pdfgen import canvas

from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

from itertools import groupby
from operator import attrgetter

from django.db.models.functions import Coalesce
from django.db.models import IntegerField

from django.conf import settings

from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
)
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.units import mm

import os



import uuid

from role.models import AffaireRoles, Decisions, Roles, MessageDefilant
from account.models import Account
from datetime import datetime, timedelta, date
from django.db.models.functions import TruncDate
from django.db.models import Count
from django.views.decorators.http import require_http_methods

from collections import defaultdict



# Create your views here.
def index(request):
    if request.user.is_authenticated:
        return backoffice(request)
    

def backoffice_data(request):
    current_year = date.today().year
    year = int(request.GET.get('year', current_year))
    today = date.today()

    # Statistiques principales
    tribunal_users = Account.objects.filter(juridiction=request.user.juridiction).count()
    visiteurs_users = Account.objects.filter(
        juridiction=request.user.juridiction,
        groups__name="Visiteur"
    ).count()

    T_roles = Roles.objects.filter(juridiction=request.user.juridiction, dateEnreg__year=year).count()
    T_roles_today = Roles.objects.filter(juridiction=request.user.juridiction, dateEnreg=today).count()

    T_roles_fond = Roles.objects.filter(juridiction=request.user.juridiction, typeAudience='Fond', dateEnreg__year=year).count()
    fond_pourcentage = round(T_roles_fond / T_roles * 100) if T_roles != 0 else 0

    T_roles_refere = Roles.objects.filter(juridiction=request.user.juridiction, typeAudience='Refere', dateEnreg__year=year).count()
    refere_pourcentage = round(T_roles_refere / T_roles * 100) if T_roles != 0 else 0

    T_affaires = AffaireRoles.objects.filter(role__juridiction=request.user.juridiction, role__dateEnreg__year=year).count()
    T_affaires_today = AffaireRoles.objects.filter(role__juridiction=request.user.juridiction, role__dateEnreg=today).count()

    T_decisions = Decisions.objects.filter(juridiction=request.user.juridiction, dateDecision__year=year).count()
    T_decisions_today = Decisions.objects.filter(juridiction=request.user.juridiction, dateDecision=today).count()

    # Graphes (Affaires sur 5 jours)
    start_date = today - timedelta(days=4)
    days = [start_date + timedelta(days=i) for i in range(5)]

    raw_stats = (
        AffaireRoles.objects
        .filter(role__juridiction=request.user.juridiction, role__dateEnreg__range=(start_date, today))
        .annotate(day=TruncDate('role__dateEnreg'))
        .values('day')
        .annotate(total=Count('id'))
    )
    stats_dict = {item['day']: item['total'] for item in raw_stats}

    chart_labels = [d.strftime('%Y-%m-%d') for d in days]
    chart_data = [stats_dict.get(d, 0) for d in days]

    # Graphes (Décisions par type)
    stats_decisions = (
        Decisions.objects
        .filter(juridiction=request.user.juridiction, dateDecision__year=year)
        .values('typeDecision')
        .annotate(total=Count('id'))
        .order_by('typeDecision')
    )
    decision_labels = [item['typeDecision'] for item in stats_decisions]
    decision_counts = [item['total'] for item in stats_decisions]

    # Messages défilants
    messages = list(
        MessageDefilant.objects.filter(actif=True)
        .order_by('-date_creation')
        .values("id", "contenu", "date_creation")
    )

    return JsonResponse({
        # Statistiques
        "tribunal_users": tribunal_users,
        "visiteurs_users": visiteurs_users,
        "T_roles": T_roles,
        "T_roles_today": T_roles_today,
        "T_roles_fond": T_roles_fond,
        "fond_pourcentage": fond_pourcentage,
        "T_roles_refere": T_roles_refere,
        "refere_pourcentage": refere_pourcentage,
        "T_affaires": T_affaires,
        "T_affaires_today": T_affaires_today,
        "T_decisions": T_decisions,
        "T_decisions_today": T_decisions_today,

        # Graphes
        "chart_labels": chart_labels,
        "chart_data": chart_data,
        "decision_labels": decision_labels,
        "decision_counts": decision_counts,

        # Messages
        "messages": messages,
    })



def backoffice(request):

    t_enrollement = Enrollement.objects.filter(dateEnrollement__year=2025)
    t_affaireRole = AffaireRoles.objects.all()

# Parcourir tous les enrôlements
    # for enr in t_enrollement:
    #     # Vérifier si un AffaireRoles correspond aux colonnes
    #     matching_roles = AffaireRoles.objects.filter(
    #         objet=enr.objet,
    #         demandeurs=enr.demandeurs,
    #         defendeurs=enr.defendeurs,
    #         role__typeAudience=enr.typeAudience  # ou role__typeAudience si relation FK
    #     )
        
    #     if matching_roles.exists():
    #         for role in matching_roles:
    #             # Mettre à jour numAffaire si différent
    #             if role.numAffaire != enr.numAffaire:
    #                 role.numAffaire = enr.numAffaire
    #                 role.idAffaire = enr.idAffaire
    #                 role.save(update_fields=["numAffaire","idAffaire"])
    #                 print(f"✅ Mis à jour: idAffaire={role.idAffaire}, numAffaire={role.numAffaire}")
    #     else:
    #         print(f"❌ Pas de correspondance pour {enr.numAffaire}")


    
        


    current_year = date.today().year
    year = int(request.GET.get('year', current_year))

    # Générer une liste d'années de 2010 à l'année courante
    available_years = list(range(2024, current_year + 1))


    today = datetime.today().date()
    start = datetime.combine(today, datetime.min.time())
    end = datetime.combine(today, datetime.max.time())

    today_roles_fond = Roles.objects.filter(juridiction=request.user.juridiction, dateEnreg=today, typeAudience='Fond')
    today_roles_refere = Roles.objects.filter(juridiction=request.user.juridiction, dateEnreg=today, typeAudience='Refere')
    today_affaires = AffaireRoles.objects.filter(role__juridiction=request.user.juridiction, role__dateEnreg=today)

    tribunal_users = Account.objects.filter(juridiction=request.user.juridiction).count()
    visiteurs_users = Account.objects.filter(
            juridiction=request.user.juridiction,
            groups__name="Visiteur"
        ).count()

    T_roles = Roles.objects.filter(juridiction=request.user.juridiction,  dateEnreg__year=year).count()
    T_roles_today = Roles.objects.filter(juridiction=request.user.juridiction, dateEnreg=today).count()

    T_roles_fond = Roles.objects.filter(juridiction=request.user.juridiction, typeAudience='Fond',  dateEnreg__year=year).count()
    fond_pourcentage = round(T_roles_fond / T_roles * 100) if T_roles != 0 else 0

    T_roles_refere = Roles.objects.filter(juridiction=request.user.juridiction, typeAudience='Refere',  dateEnreg__year=year).count()
    refere_pourcentage = round(T_roles_refere / T_roles * 100) if T_roles != 0 else 0


    T_enrollements = Enrollement.objects.filter(juridiction=request.user.juridiction).count()

    T_affaires = AffaireRoles.objects.filter(role__juridiction=request.user.juridiction,  role__dateEnreg__year=year).count()
    T_affaires_today = AffaireRoles.objects.filter(role__juridiction=request.user.juridiction, role__dateEnreg=today).count()

    T_decisions = Decisions.objects.filter(juridiction=request.user.juridiction, dateDecision__year=year).count()
    T_decisions_today = Decisions.objects.filter(juridiction=request.user.juridiction, dateDecision=today).count()

    T_affaires_sp = AffaireRoles.objects.filter(role__juridiction=request.user.juridiction, role__section='Section Présidentielle',  role__dateEnreg__year=year).count()
    president_sp = AffaireRoles.objects.filter(role__juridiction=request.user.juridiction, role__section='Section Présidentielle').last()

    T_affaires_s1 = AffaireRoles.objects.filter(role__juridiction=request.user.juridiction, role__section='Premiere-Section',  role__dateEnreg__year=year).count()
    president_s1 = AffaireRoles.objects.filter(role__juridiction=request.user.juridiction, role__section='Premiere-Section').last()

    T_affaires_s2 = AffaireRoles.objects.filter(role__juridiction=request.user.juridiction, role__section='Deuxieme-Section',  role__dateEnreg__year=year).count()
    president_s2 = AffaireRoles.objects.filter(role__juridiction=request.user.juridiction, role__section='Deuxieme-Section').last()

    T_affaires_s3 = AffaireRoles.objects.filter(role__juridiction=request.user.juridiction, role__section='Troisieme-Section',  role__dateEnreg__year=year).count()
    president_s3 = AffaireRoles.objects.filter(role__juridiction=request.user.juridiction, role__section='Troisieme-Section').last()

    T_affaires_s4 = AffaireRoles.objects.filter(role__juridiction=request.user.juridiction, role__section='Quatrieme-Section',  role__dateEnreg__year=year).count()
    president_s4 = AffaireRoles.objects.filter(role__juridiction=request.user.juridiction, role__section='Quatrieme-Section').last()

    T_affaires_s5 = AffaireRoles.objects.filter(role__juridiction=request.user.juridiction, role__section='Cinquieme-Section',  role__dateEnreg__year=year).count()
    president_s5 = AffaireRoles.objects.filter(role__juridiction=request.user.juridiction, role__section='Cinquieme-Section').last()

    # Graphes 1
    today = date.today()
    start_date = today - timedelta(days=4)

    # Tous les jours de l'intervalle (pour inclure les jours sans enregistrement)
    days = [start_date + timedelta(days=i) for i in range(5)]

    # Stats depuis la base
    raw_stats = (
        AffaireRoles.objects
        .filter(
            role__juridiction=request.user.juridiction,
            role__dateEnreg__range=(start_date, today)
        )
        .annotate(day=TruncDate('role__dateEnreg'))
        .values('day')
        .annotate(total=Count('id'))
    )


    # Dictionnaire jour => total
    stats_dict = {item['day']: item['total'] for item in raw_stats}

    # Données formatées pour le graphique
    labels = [d.strftime('%Y-%m-%d') for d in days]  # ou '%Y-%m-%d' selon ton format préféré
    data = [stats_dict.get(d, 0) for d in days]  # 0 si pas d'enregistrement ce jour-là


     # Graphes 2
    stats_decisions = (
        Decisions.objects
        .filter(
            juridiction=request.user.juridiction,
            dateDecision__year=year
        )
        .values('typeDecision')
        .annotate(total=Count('id'))
        .order_by('typeDecision')  # optionnel
    )

    # Conversion en deux listes
    decision_labels = [item['typeDecision'] for item in stats_decisions]
    decision_counts = [item['total'] for item in stats_decisions]


    messages = MessageDefilant.objects.filter(actif=True).order_by('-date_creation')


    context = {
        'today':today,
        'selected_year': year,
        'available_years': available_years,
        'today_roles_fond':today_roles_fond,
        'today_roles_refere':today_roles_refere,
        'today_affaires':today_affaires,
        'T_roles':T_roles,
        'T_enrollements':T_enrollements,
        'T_roles_today':T_roles_today,
        'T_affaires':T_affaires,
        'T_affaires_today':T_affaires_today,
        'T_decisions':T_decisions,
        'T_decisions_today':T_decisions_today,
        'T_roles_fond':T_roles_fond,
        'T_roles_refere':T_roles_refere,
        'fond_pourcentage':fond_pourcentage,
        'refere_pourcentage':refere_pourcentage,
        'T_affaires_sp':T_affaires_sp,
        'president_sp':president_sp,
        'T_affaires_s1':T_affaires_s1,
        'president_s1':president_s1,
        'T_affaires_s2':T_affaires_s2,
        'president_s2':president_s2,
        'T_affaires_s3':T_affaires_s3,
        'president_s3':president_s3,
        'T_affaires_s4':T_affaires_s4,
        'president_s4':president_s4,
        'T_affaires_s5':T_affaires_s5,
        'president_s5':president_s5,
        'chart_labels': labels,
        'chart_data': data,
        'decision_labels': decision_labels,
        'decision_counts': decision_counts,
        'tribunal_users': tribunal_users,
        'visiteurs_users': visiteurs_users,
        'messages': messages,
        
    }

    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return render(request, 'role/partial_accueil.html', context)  # HTML partiel

    return render(request, 'role/index-backoffice.html', context)



def colorize_found(query, text):
    colored_text = re.sub(r'(' + re.escape(query) + r')', r'<span style="color:red;">\1</span>', text, flags=re.IGNORECASE)
    return mark_safe(colored_text)

def listRole(request):
    # Récupération des paramètres GET
    current_year = date.today().year
    year = int(request.GET.get('year', current_year))
    query = request.GET.get('q', '').strip()
    type_audience = request.GET.get('typeAudience', '').strip()
    section = request.GET.get('section', '').strip()
    selected_date = request.GET.get('date', '').strip()

    # Années disponibles pour le filtre
    available_years = list(range(2024, current_year + 1))

    # Base queryset
    roles = Roles.objects.all().order_by('dateEnreg')

    # Filtrage par année
    if year:
        roles = roles.filter(dateEnreg__year=year)

    # Filtrage par typeAudience
    if type_audience:
        roles = roles.filter(typeAudience=type_audience)

    # Filtrage par section
    if section:
        roles = roles.filter(section=section)

    # Filtrage par date précise
    if selected_date:
        try:
            parsed_date = datetime.strptime(selected_date, "%Y-%m-%d").date()
            roles = roles.filter(dateEnreg=parsed_date)
        except ValueError:
            pass  # ignore si la date est mal formée

    # Filtrage par recherche globale
    if query:
        roles = roles.filter(
            Q(typeAudience__icontains=query) |
            Q(section__icontains=query) |
            Q(president__icontains=query) |
            Q(greffier__icontains=query) |
            Q(dateEnreg__icontains=query)
        )

    # Pagination
    paginator = Paginator(roles, 20)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    # Options pour les filtres
    type_audiences = (
        ("Refere", "Référé"),
        ("Fond", "Fond"),
    )

    sections = (
        ("Premiere-Section", "Première Section"),
        ("Deuxieme-Section", "Deuxième Section"),
        ("Troisieme-Section", "Troisième Section"),
        ("Quatrieme-Section", "Quatrième Section"),
        ("Cinquieme-Section", "Cinquième Section"),
        ("Section-Presidentielle", "Section Présidentielle"),
    )

    context = {
        'roles': roles,  # utile si accès direct nécessaire
        'page_obj': page_obj,
        'selected_year': year,
        'available_years': available_years,
        'query': query,
        'type_audiences': type_audiences,
        'sections': sections,
        'selected_type_audience': type_audience,
        'selected_section': section,
        'selected_date': selected_date
    }

    return render(request, 'role/gestion-roles.html', context)

def listAffaire(request):
    # Paramètres GET
    current_year = date.today().year
    year = int(request.GET.get('year', current_year))
    query = request.GET.get('q', '').strip()
    type_audience = request.GET.get('typeAudience', '').strip()
    section = request.GET.get('section', '').strip()
    selected_date = request.GET.get('date', '').strip()

    # Années disponibles
    available_years = list(range(2024, current_year + 1))

    # Base queryset
    affaires = AffaireRoles.objects.select_related('role').all().order_by('role__dateEnreg')

    # Filtrage par année
    if year:
        affaires = affaires.filter(role__dateEnreg__year=year)

    # Filtrage par typeAudience
    if type_audience:
        affaires = affaires.filter(role__typeAudience=type_audience)

    # Filtrage par section
    if section:
        affaires = affaires.filter(role__section=section)

    # Filtrage par date précise (role__dateEnreg ou role__dateAudience si besoin)
    if selected_date:
        try:
            parsed_date = datetime.strptime(selected_date, "%Y-%m-%d").date()
            affaires = affaires.filter(role__dateEnreg=parsed_date)
        except ValueError:
            pass  # Ignore si la date est mal formée

    # Recherche texte
    if query:
        affaires = affaires.filter(
            Q(numRg__icontains=query) |
            Q(numAffaire__icontains=query) |
            Q(demandeurs__icontains=query) |
            Q(defendeurs__icontains=query) |
            Q(objet__icontains=query)
        )

    # Pagination
    paginator = Paginator(affaires, 20)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    # Options pour les filtres
    type_audiences = (
        ("Refere", "Référé"),
        ("Fond", "Fond"),
    )

    sections = (
        ("Premiere-Section", "Première Section"),
        ("Deuxieme-Section", "Deuxième Section"),
        ("Troisieme-Section", "Troisième Section"),
        ("Quatrieme-Section", "Quatrième Section"),
        ("Cinquieme-Section", "Cinquième Section"),
        ("Section-Presidentielle", "Section Présidentielle"),
    )

    # Récupérer toutes les décisions liées aux affaires affichées
    affaire_ids = [a.id for a in page_obj]
    decisions = Decisions.objects.filter(affaire_id__in=affaire_ids).select_related("affaire")

    # Indexer les décisions par affaire_id
    decisions_map = {d.affaire_id: d.decision for d in decisions}

    # Ajouter la décision correspondante à chaque affaire
    for affaire in page_obj:
        affaire.decision = decisions_map.get(affaire.id, "-")

    context = {
        'page_obj': page_obj,
        'selected_year': year,
        'available_years': available_years,
        'query': query,
        'type_audiences': type_audiences,
        'sections': sections,
        'selected_type_audience': type_audience,
        'selected_section': section,
        'selected_date': selected_date,
    }

    return render(request, 'role/gestion-affaires.html', context)


def listEnrollement(request):
    if request.user.groups.filter(name='Greffe').exists():
        enrollements = Enrollement.objects.filter(juridiction=request.user.juridiction_id)
        return render(request, 'role/gestion-enrollements.html',{'enrollements':enrollements})
    else:
        juridictions = Juridictions.objects.all()

        # Nombre d'objets par page
        objets_par_page = 12

        paginator = Paginator(juridictions, objets_par_page)

        # Récupérez le numéro de page à partir de la requête GET
        page_number = request.GET.get('page')
        
        # Obtenez les objets pour la page demandée
        juridictions = paginator.get_page(page_number)

        return render(request, 'role/registres-enrollements.html',{'juridictions':juridictions})


def listEnrollementForAdmin(request):
    # Récupération des paramètres GET
    current_year = date.today().year
    year = int(request.GET.get('year', current_year))
    query = request.GET.get('q', '').strip()
    type_audience = request.GET.get('typeAudience', '').strip()
    section = request.GET.get('section', '').strip()
    selected_date = request.GET.get('date', '').strip()
    available_years = list(range(2024, current_year + 1))

    # Base queryset
    enrollements = Enrollement.objects.all().order_by('dateEnrollement')

    # Filtrage par année
    if year:
        enrollements = enrollements.filter(dateEnrollement__year=year)

    # Filtrage par typeAudience
    if type_audience:
        enrollements = enrollements.filter(typeAudience=type_audience)

    # Filtrage par section
    if section:
        enrollements = enrollements.filter(section=section)

    # Filtrage par date précise (dateEnrollement et dateAudience)
    if selected_date:
        try:
            parsed_date = datetime.strptime(selected_date, "%Y-%m-%d").date()
            enrollements = enrollements.filter(
                Q(dateEnrollement=parsed_date)
            )
        except ValueError:
            # Si la date est mal formée, on ignore le filtre
            pass

    # Filtrage par recherche globale
    if query:
        enrollements = enrollements.filter(
            Q(numRg__icontains=query) |
            Q(numAffaire__icontains=query) |
            Q(typeAudience__icontains=query) |
            Q(section__icontains=query) |
            Q(objet__icontains=query) |
            Q(demandeurs__icontains=query) |
            Q(defendeurs__icontains=query) |
            Q(dateEnrollement__icontains=query) |
            Q(dateAudience__icontains=query)
        )

    # Pagination
    paginator = Paginator(enrollements, 20)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    # Choices pour le template
    sections = (
        ("Premiere-Section", "Prémière Section"),
        ("Deuxieme-Section", "Deuxième Section"),
        ("Troisieme-Section", "Troisième Section"),
        ("Quatrieme-Section", "Quatrième Section"),
        ("Cinquieme-Section", "Cinquième Section"),
        ("Section-Presidentielle", "Section Présidentielle"),
    )

    type_audiences = (
        ("Refere", "Référé"),
        ("Fond", "Fond"),
    )

    return render(request, 'role/gestion-enrollements.html', {
        'page_obj': page_obj,
        'available_years': available_years,
        'selected_year': year,
        'sections': sections,
        'type_audiences': type_audiences,
        'selected_type_audience': type_audience,
        'selected_section': section,
        'selected_date': selected_date,
        'query': query,
    })


def edit_affaire(request, idAffaire):

    enrollement = get_object_or_404(Enrollement, id=idAffaire)

    numAffaire = request.POST.get('numAffaire', '').strip()
    affaireRole = AffaireRoles.objects.filter(numAffaire=numAffaire)
  
    if affaireRole :

        for aff in affaireRole:
            # Mise à jour dans affaire au role
            aff.demandeurs = request.POST.get('demandeurs', '').strip()
            aff.defendeurs = request.POST.get('defendeurs', '').strip()
            aff.objet = request.POST.get('objet', '').strip()
            aff.save()

    else :

        # Historique modification
        old = Enrollement.objects.get(pk=enrollement.id)
        EnrollementHistory.objects.create(
            original=old,
            numOrdre=old.numOrdre,
            numRg=old.numRg,
            numAffaire=old.numAffaire,
            objet=old.objet,
            decision=old.decision,
            demandeurs=old.demandeurs,
            defendeurs=old.defendeurs,
            dateEnrollement=old.dateEnrollement,
            dateAudience=old.dateAudience,
            juridiction=old.juridiction,
            typeAudience=old.typeAudience,
            section=old.section,
            statut=old.statut,
            motifAnnulation=old.motifAnnulation,
            modified_by=request.user
        )
          # Mise à jour dans enrollement
        enrollement.typeAudience = request.POST.get('typeAudience', '').strip()
        enrollement.dateEnrollement = request.POST.get('dateEnrollement')
        enrollement.dateAudience = request.POST.get('dateAudience')
        enrollement.demandeurs = request.POST.get('demandeurs', '').strip()
        enrollement.defendeurs = request.POST.get('defendeurs', '').strip()
        enrollement.objet = request.POST.get('objet', '').strip()
        enrollement.statut = 'Modifier'
        enrollement.save()


    messages.success(request, 'Affaire modifiée avec succès !')

    return redirect('role.enrollementForAdmin')
   

def createRole(request):
    juridictions = Juridictions.objects.filter(id=request.user.juridiction_id)
    form = RoleForm(request.POST or None)

    enrollFormset = modelformset_factory(
        Enrollement, form=EnrollementForm, extra=0, exclude=['id']
    )
    formset = enrollFormset(
        request.POST or None,
        queryset=Enrollement.objects.filter(
            juridiction_id=request.user.juridiction_id
        ).exclude(statut="Annuler")
    )


    if request.method == 'POST':
        if form.is_valid():
            try:
                with transaction.atomic():
                    juridiction_id = request.POST.get('juridiction_id')
                    juridiction = Juridictions.objects.get(pk=juridiction_id)

                    # Création du rôle
                    role = form.save(commit=False)
                    role.juridiction = juridiction
                    role.created_by = request.user
                    role.save()

                    if formset.is_valid():
                        for affaire_form in formset:
                            if not affaire_form.cleaned_data:
                                continue

                            numAffaire = affaire_form.cleaned_data.get('numAffaire')

                            # Si pas de numAffaire → créer un Enrollement avant de créer l’AffaireRoles
                            if not numAffaire:
                                enr = affaire_form.save(commit=False)
                                enr.juridiction = juridiction
                                enr.typeAudience = role.typeAudience
                                enr.section = role.section
                                enr.dateEnrollement = date(2024, 12, 31)  
                                enr.dateAudience = role.dateEnreg  
                                enr.created_by = request.user
                                enr.save()

                                # Génération du numéro d’affaire
                                date_str = enr.dateEnrollement.strftime("%d%m%y")
                                tribunal = 'GNTC'
                                initial = 'R' if enr.typeAudience == 'Refere' else 'F'
                                enr.numAffaire = f"JUD{initial}{date_str}{tribunal}{enr.id}"
                                enr.save()

                                numAffaire = enr.numAffaire
                                id_affaire = enr.idAffaire
                            else:
                                id_affaire = affaire_form.cleaned_data.get('idAffaire') or uuid.uuid4()

                            # Création de l'affaire liée au rôle
                            affaireEnroller = AffaireRoles(
                                role=role,
                                idAffaire=id_affaire,
                                numOrdre=affaire_form.cleaned_data.get('numOrdre'),
                                numRg=affaire_form.cleaned_data.get('numRg'),
                                numAffaire=numAffaire,
                                objet=affaire_form.cleaned_data.get('objet'),
                                mandatDepot=affaire_form.cleaned_data.get('mandatDepot'),
                                detention=affaire_form.cleaned_data.get('detention'),
                                prevention=affaire_form.cleaned_data.get('prevention'),
                                natureInfraction=affaire_form.cleaned_data.get('natureInfraction'),
                                decision=affaire_form.cleaned_data.get('decision'),
                                prevenus=affaire_form.cleaned_data.get('prevenus'),
                                demandeurs=affaire_form.cleaned_data.get('demandeurs'),
                                defendeurs=affaire_form.cleaned_data.get('defendeurs'),
                                appelants=affaire_form.cleaned_data.get('appelants'),
                                intimes=affaire_form.cleaned_data.get('intimes'),
                                partieCiviles=affaire_form.cleaned_data.get('partieCiviles'),
                                civileResponsables=affaire_form.cleaned_data.get('civileResponsables'),
                                created_by=request.user,
                            )
                            affaireEnroller.save()

                        messages.success(request, "Rôle et affaires enregistrés avec succès !")
                        return redirect("role.liste")

                    else:
                        messages.error(request, f"Erreur Formset : {formset.errors}")
                        return redirect("role.create")

            except IntegrityError as e:
                messages.error(request, f"Erreur d'intégrité : {e}")
                return redirect("role.create")

    context = {
        'juridictions': juridictions,
        'form': form,
        'formset': formset,
    }
    return render(request, 'role/new-role.html', context)


def valide_role(request, pk):
    role=Roles.objects.get(pk=pk)
    role.statut='Valider'
    role.save()
    messages.success(request, 'Rôle validé avec succès !')
    return redirect('role.detail', pk=role.idRole)


def createEnrollement(request):

    juridictions = Juridictions.objects.filter(id=request.user.juridiction_id)

    context = {}
    form = RoleForm(request.POST or None)
    EnrollementFormset = modelformset_factory(Enrollement, form=EnrollementForm, extra=0)
    formset = EnrollementFormset(request.POST or None, queryset=Enrollement.objects.none())

  
    if request.method == 'POST':
       
        if form.is_valid() and formset.is_valid():
           
            try:
                with transaction.atomic():
                    juridiction_id = request.POST.get('juridiction_id')
                    typeAudience = request.POST.get('typeAudience')
                    section = 'Section-Presidentielle'
                    juridiction = Juridictions.objects.get(pk=juridiction_id)
                    
                    for affaire_form in formset:
                        affaire = affaire_form.save(commit=False)
                        affaire.juridiction = juridiction
                        affaire.typeAudience = typeAudience
                        affaire.section = section
                        affaire.created_by = request.user
                         # Vérification si l'affaire existe déjà dans la BD en fonction de certains champs
                        try:
                            affaire_existe = Enrollement.objects.filter(
                                numOrdre=affaire.numOrdre, 
                                numRg=affaire.numRg, 
                                juridiction=juridiction,
                                typeAudience=typeAudience,
                                section=section,
                                dateAudience=affaire.dateAudience
                            ).exists()

                            if affaire_existe:
                                # Si l'affaire existe déjà, ne pas l'enregistrer et passer à la suivante
                                # messages.warning(request, f"L'affaire avec le numéro d'ordre {affaire.numOrdre} existe déjà et n'a pas été enregistrée.")
                                continue
                            else:
                                # Si l'affaire n'existe pas, on l'enregistre
                                affaire.save()

                                date_str = affaire.dateEnrollement.strftime("%d%m%y")
                                r_f = affaire.typeAudience
                                tribunal = 'GNTC'
                                if r_f == 'Refere':
                                    initial = 'R'
                                else:
                                    initial = 'F'

                                affaire.numAffaire = f"JUD{initial}{date_str}{tribunal}{affaire.id}"

                                affaire.save()  # Sauvegarde finale avec numRg mis à jour
                                
                        except Exception as e:
                            messages.error(request, f"Erreur lors de l'enregistrement de l'affaire : {e}")
                            return redirect('role.createEnrollement')
                        
                    messages.success(request, 'Affaire(s) enrollée(s) avec succès !')
                    return redirect('role.enrollementForAdmin')

            except IntegrityError as e:
                messages.error(request, f"Erreur d'intégrité : {e}")
                return redirect('role.createEnrollement')
            except Exception as e:
                messages.error(request, f"Une erreur est survenue : {e}")
                return redirect('role.createEnrollement')
        else:
            # Si form ou formset ne sont pas valides, afficher les erreurs
            if form.errors:
                messages.error(request, f"Erreurs dans le formulaire principal : {form.errors}")
            if formset.errors:
                messages.error(request, f"Erreurs dans le formset : {formset.errors}")    

    is_chef = request.user.groups.filter(name='Chef').exists()

    context = {
        'juridictions':juridictions,
        'form':form,
        'formset':formset,
        'is_chef':is_chef,
    }        
    return render(request, 'role/new-enrollement.html',context)

@require_http_methods(["POST"])
def cancel_affaire(request, id):
    affaire = get_object_or_404(Enrollement, id=id)
    numAffaire = affaire.numAffaire
    affaireRole = AffaireRoles.objects.filter(numAffaire=numAffaire).first()
    if affaireRole :
         messages.error(request, 'Cette affaire est déjà au rôle !')
    else :
        affaire.statut = 'Annuler'
        affaire.motifAnnulation = request.POST.get('motifAnnulation')
        affaire.save()

        messages.success(request, 'Affaire annulée avec succès !')
    
    return redirect('role.enrollementForAdmin')  # ou autre URL de redirection

def roleDetail(request, pk):
    search_query = request.GET.get('q', '').strip()
    role = Roles.objects.filter(idRole=pk).first()

    if not role:
        return HttpResponse("Rôle non trouvé", status=404)

    # Sous-requête : compter toutes les décisions avec le même numAffaire
    decisions_count_subquery = (
        Decisions.objects.filter(numAffaire=OuterRef('numAffaire'))
        .values('numAffaire')
        .annotate(total=Count('id'))
        .values('total')[:1]
    )

    # Annoter chaque affaire avec nb_decisions global et catégorie
    affaires = AffaireRoles.objects.filter(role=role).annotate(
        nb_decisions=Coalesce(
            Subquery(decisions_count_subquery, output_field=IntegerField()), 
            0
        ),
        categorie=Case(
            When(nb_decisions__lt=2, then=Value('Nouvelles Affaires')),
            When(nb_decisions__gte=2, then=Value('Affaires Encours')),
            output_field=CharField(),
        )
    ).order_by('categorie', 'numOrdre')

    # Recherche
    if search_query:
        affaires = affaires.filter(Q(objet__icontains=search_query)|Q(demandeurs__icontains=search_query)| Q(defendeurs__icontains=search_query)|Q(objet__icontains=search_query))

    # Pagination unique
    paginator = Paginator(affaires, 10)  # 10 affaires par page
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    # Grouper par catégorie pour le template
    sorted_affaires = sorted(page_obj.object_list, key=attrgetter('categorie', 'numOrdre'))
    grouped_affaires = []
    for categorie, items in groupby(sorted_affaires, key=attrgetter('categorie')):
        grouped_affaires.append({
            'grouper': categorie,
            'items': list(items)
        })

    # Infos utilisateur et juridiction
    juridiction = Juridictions.objects.filter(id=role.juridiction_id).first()
    is_chef = request.user.groups.filter(name='Chef').exists()
    affaireSuivis = SuivreAffaire.objects.filter(account=request.user) if request.user.is_authenticated else SuivreAffaire.objects.none()

    context = {
        'role': role,
        'grouped_affaires': grouped_affaires,
        'page_obj': page_obj,
        'is_chef': is_chef,
        'affaireSuivis': affaireSuivis
    }

    # Choix du template selon juridiction et type d'audience
    if juridiction and juridiction.name == 'Tribunal de Commerce de Conakry' and role.typeAudience == 'Fond':
        return render(request, 'role/details/tc-fond-detail.html', context)
    elif juridiction and juridiction.name == 'Tribunal de Commerce de Conakry' and role.typeAudience == 'Refere':
        return render(request, 'role/details/tc-refere-detail.html', context)
    else:
        return HttpResponse("Template non disponible pour cette juridiction/type d'audience")
    

def detailAffaire(request, idAffaire):


    type_section = (

        ("Premiere-Section", "Prémière Section"),
        ("Deuxieme-Section", "Deuxième Section"),
        ("Troisieme-Section", "Troisième Section"),
        ("Quatrieme-Section", "Quatrième Section"),
        ("Cinquieme-Section", "Cinquième Section"),
        ("Section-Presidentielle", "Section Présidentielle"),
    )

    type_decisions = (
        ("Renvoi", "Renvoi"),
        ("Mise-en-delibere", "Mise en délibéré"),
        ("Delibere-proroge", "Délibéré prorogé"),
        ("Vide-du-delibere", "Vidé du délibéré"),
        ("Radiation", "Radiation"),
        ("Renvoi-sine-die", "Renvoi sine die"),
        ("Affectation", "Affectation"),
        ("Autre", "Autre"),
    )

    affaire = AffaireRoles.objects.filter(idAffaire=idAffaire).first()
    print(idAffaire)
    decisions = Decisions.objects.select_related('affaire').filter(
        affaire__objet=affaire.objet,
        affaire__demandeurs=affaire.demandeurs,
        affaire__defendeurs=affaire.defendeurs,
        affaire__mandatDepot=affaire.mandatDepot,
        affaire__detention=affaire.detention,
        affaire__prevention=affaire.prevention,
        affaire__natureInfraction=affaire.natureInfraction,
        affaire__prevenus=affaire.prevenus,
        affaire__appelants=affaire.appelants,
        affaire__intimes=affaire.intimes,
        affaire__partieCiviles=affaire.partieCiviles,
        affaire__civileResponsables=affaire.civileResponsables
    )
    affaireRole = AffaireRoles.objects.select_related('role__juridiction').get(id=affaire.id)
    affaireEnroller = Enrollement.objects.filter(idAffaire=idAffaire).first()


    is_suivi = SuivreAffaire.objects.filter(affaire=affaire,juridiction=affaireRole.role.juridiction,account=request.user)
    is_greffe = request.user.groups.filter(name='Greffe').exists()
    juridiction = Juridictions.objects.filter(id=request.user.juridiction_id).first()

    if request.method == 'POST':
        form = DecisionsForm(request.POST)
        newSection = request.POST.get('section')
        juridiction_id = request.POST.get('juridiction')
        typeAudience = request.POST.get('typeAudience')

        if form.is_valid():
            decision = form.save(commit=False)
            decision.affaire = affaire
            decision.numAffaire = affaire.numAffaire
            decision.juridiction = juridiction
            decision.created_by = request.user
            decision.dateDecision = affaire.role.dateEnreg

            if newSection:
                decision.section = newSection
            else:
                decision.section = affaire.role.section
          
            form.save()



        
            messages.success(request, 'Décision ajoutée avec succès !')
            return redirect(request.META.get('HTTP_REFERER', '/'))  
    else:
        form = DecisionsForm()

    

    context = {
        'affaire':affaire,
        'affaireEnroller':affaireEnroller,
        'decisions':decisions,
        'is_greffe':is_greffe,
        'is_suivi':is_suivi,
        'type_section':type_section,
        'type_decisions': type_decisions,
        'form': form
    }

    # Formater l'URL avec l'ID dynamique
    url = f'/role/affaires/details/{idAffaire}'

    # Effectuer la mise à jour
    Notification.objects.filter(
        Q(recipient=request.user) & 
        Q(url=url) & 
        Q(is_read=False)
    ).update(is_read=True)
    return render(request, 'role/detail-affaire.html',context)
  

def fetchForm(request, selectedJuridiction, selectedType, dateRole, selectedSection):

    juridiction = Juridictions.objects.filter(id=selectedJuridiction).first()

    affaireEnrollers = Enrollement.objects.filter(
        juridiction=juridiction, typeAudience=selectedType, dateAudience=dateRole, section=selectedSection, statut='Creer'
    )
    decisionsRenvoyers = Decisions.objects.filter(
        Q(prochaineAudience=dateRole) &
        Q(juridiction=juridiction) &
        Q(affaire__role__typeAudience=selectedType) &
        Q(section=selectedSection) &
       (Q(typeDecision='Renvoi') | Q(typeDecision='Mise-en-delibere') | Q(typeDecision='Delibere-proroge') | Q(typeDecision='Affectation'))
    ).select_related('affaire')
    
    for decision in decisionsRenvoyers:
        print("ID de la décision:", decision.idDecision)
        print("Juridiction:", decision.juridiction)
        print("Type de décision:", decision.typeDecision)
        print("Objet:", decision.objet)
        print("Président:", decision.president)
        print("Greffier:", decision.greffier)
        print("Date de la décision:", decision.dateDecision)
        print("Prochaine audience:", decision.prochaineAudience)
        print("Affaire liée:", decision.affaire)
        if decision.affaire:  # Pour afficher des détails de l'affaire liée
            print("ID de l'affaire:", decision.affaire.idAffaire)
            print("Objet de l'affaire:", decision.affaire.objet)
        print("--------")


    verifRole = Roles.objects.filter(
        juridiction=juridiction, typeAudience=selectedType, dateEnreg=dateRole, section=selectedSection
    )
    message = ''

    default_data = []

    # Vérifiez si un rôle existe déjà
    if verifRole.exists():
        message = 'Le rôle pour cette date a déjà été enregistré !'
        return render(request, 'role/roleForms/message_role_exist.html', {'message': message})
    else:

        # Ajoutez les données de `decisionsRenvoyers` à `default_data`
        for d in decisionsRenvoyers:
            default_data.append({
                'typeDecision': d.typeDecision,
                'decision': d.decision,
                'prochaineAudience': d.prochaineAudience,
                'president': d.president,
                'greffier': d.greffier,
                'dateDecision': d.dateDecision,
                'numRg': d.affaire.numRg,
                'demandeurs': d.affaire.demandeurs,
                'defendeurs': d.affaire.defendeurs,
                'numAffaire': d.affaire.numAffaire,
                'objet': d.affaire.objet,
                'idAffaire': d.affaire.idAffaire if d.affaire else None,  # Exemple d'accès à `AffaireRoles`
            })

        # Ajoutez les données d'`affaireEnrollers` à `default_data`
        for a in affaireEnrollers:
            default_data.append({
                'numOrdre': a.numOrdre,
                'idAffaire': a.idAffaire,
                'numRg': a.numRg,
                'numAffaire': a.numAffaire,
                'demandeurs': a.demandeurs,
                'defendeurs': a.defendeurs,
                'objet': a.objet,
                'dateEnrollement': a.dateEnrollement,
                'dateAudience': a.dateAudience
            })

       
        # Initialisez le formset avec `default_data`
        enrollementFormset = modelformset_factory(Enrollement, form=EnrollementForm, extra=len(default_data))
        formset = enrollementFormset(request.POST or None, queryset=Enrollement.objects.none(), initial=default_data)

    # Création du formulaire principal
    form = RoleForm(request.POST or None)

    president = ''
    greffier = ''

    if selectedSection == 'Premiere-Section':
        president = 'M. Fulber Aimé SAGNO'
        greffier = 'Mme. Hawanatou Djoubar SOUMAH'

    elif selectedSection == 'Deuxieme-Section':
        president = 'M. Mamoudou CAMARA'
        greffier = 'Mme Maïmouna DIALLO'
    
    elif selectedSection == 'Troisieme-Section':
        president = 'M. Kaman Magloire Théophile KOUADIO'
        greffier = 'M. Sekou Mohamed CAMARA'
    
    elif selectedSection == 'Quatrieme-Section':
        president = 'M. Mamadou KABA'
        greffier = 'Mme Aminata DOUNO'
    
    elif selectedSection == 'Cinquieme-Section':
        president = 'M. Alpha Oumar CAMARA'
        greffier = 'Mme Béatrice Tounkara'

    elif selectedSection == 'Section-Presidentielle':
        president = 'M. Sekou Kande'
        greffier = 'M. Abdoulaye Yarie Soumah'


    # Contexte pour les templates
    context = {
        'formset': formset,
        'form': form,
        'message': message,
        'default_data': default_data,
        'affaireEnrollers': affaireEnrollers,
        'decisionsRenvoyers': decisionsRenvoyers,
        'selectedJuridiction': selectedJuridiction,
        'selectedType': selectedType,
        'dateRole': dateRole,
        'selectedSection': selectedSection,
        'president': president,
        'greffier': greffier,
    }

    # Chargement de différents templates selon `juridiction` et `selectedType`
    if juridiction.name == 'Tribunal de Commerce de Conakry' and selectedType == 'Fond':
        return render(request, 'role/roleForms/tc-fond.html', context)
    elif juridiction.name == 'Tribunal de Commerce de Conakry' and selectedType == 'Refere':
        return render(request, 'role/roleForms/tc-refere.html', context)
    else:
        return HttpResponse()


def fetchFormEnrollement(request, selectedJuridiction,selectedType):
    juridiction = Juridictions.objects.filter(id=selectedJuridiction).first()
    enrollementFormset = modelformset_factory(Enrollement, form=EnrollementForm, extra=1)
    formset = enrollementFormset(request.POST or None, queryset=Enrollement.objects.none())
    form = EnrollementForm(request.POST or None)

    context = {
        'formset':formset,
        'form':form,
    }
    if juridiction.name=='Tribunal de Commerce de Conakry' and selectedType=='Fond':
        return render(request, 'role/enrollementForms/tc-fond.html',context)
    elif juridiction.name=='Tribunal de Commerce de Conakry' and selectedType=='Refere':
        return render(request, 'role/enrollementForms/tc-refere.html',context)
    else:
        return HttpResponse()

def download_pdf(request):
    # Récupérer le contenu HTML de la requête POST
    html_content = request.POST.get('html_content', '')

    # Convertir le HTML en PDF avec weasyprint
    pdf_file = html(string=html_content).write_pdf()

    # Créer une réponse avec le PDF
    response = HttpResponse(pdf_file, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="contenu.pdf"'

    return response


def updateRole(request):

    if request.method == 'POST':
        if request.POST.get('idAffaire'):        
            idAffaire = request.POST.get('idAffaire')
            # Table affaire role
            obj = AffaireRoles.objects.filter(idAffaire=idAffaire).first()

            if request.POST.get('demandeurs'): 
                obj.demandeurs = request.POST.get('demandeurs')
            if request.POST.get('defendeurs'): 
                obj.defendeurs = request.POST.get('defendeurs')
            if request.POST.get('objet'): 
                obj.objet = request.POST.get('objet')
            if request.POST.get('decision'): 
                obj.decision = request.POST.get('decision')
        
            obj.save()

            # Table enrollement
            obj2 = Enrollement.objects.filter(idAffaire=idAffaire).first()

            if obj2 is not None:
                if request.POST.get('demandeurs'): 
                    obj2.demandeurs = request.POST.get('demandeurs')
                if request.POST.get('defendeurs'): 
                    obj2.defendeurs = request.POST.get('defendeurs')
                if request.POST.get('objet'): 
                    obj2.objet = request.POST.get('objet')
                if request.POST.get('decision'): 
                    obj2.decision = request.POST.get('decision')

                obj2.save()
            
        else:
            idRole = request.POST.get('idRole')
            obj = Roles.objects.filter(id=idRole).first()

            if request.POST.get('dateEnreg'): 
                obj.dateEnreg = request.POST.get('dateEnreg')
            if request.POST.get('president'): 
                obj.president = request.POST.get('president')
            if request.POST.get('juge'): 
                obj.juge = request.POST.get('juge')
            if request.POST.get('greffier'):
                obj.greffier = request.POST.get('greffier')
            if request.POST.get('assesseur'):
                obj.assesseur = request.POST.get('assesseur')
            if request.POST.get('assesseur1'):
                obj.assesseur1 = request.POST.get('assesseur1')
            if request.POST.get('assesseur2'):
                obj.assesseur2 = request.POST.get('assesseur2')
            if request.POST.get('conseillers'):
                obj.conseillers = request.POST.get('conseillers')
            if request.POST.get('ministerePublic'):
                obj.ministerePublic = request.POST.get('ministerePublic')
            if request.POST.get('typeAudience'):
                obj.typeAudience = request.POST.get('typeAudience')
            if request.POST.get('dateEnreg'):
                obj.dateEnreg = request.POST.get('dateEnreg')
            if request.POST.get('procureurMilitaire'):
                obj.procureurMilitaire = request.POST.get('procureurMilitaire')
            if request.POST.get('subtituts'):
                obj.subtituts = request.POST.get('subtituts')
            

            obj.save()
            
    return redirect(request.META.get('HTTP_REFERER', '/'))

def deleteRole(request):
    role = get_object_or_404(Roles, id=request.POST.get('idRole'))
    role.delete()
    messages.success(request, 'Rôle supprimé avec succès !')
    return redirect('role.liste')

def deleteDecision(request):
    decision = get_object_or_404(Decisions, id=request.POST.get('idDecision'))
    decision.delete()
    messages.success(request, 'Décision supprimée avec succès !')
    idAffaire=request.POST.get('idAffaire')
    return redirect('affaires.details' , idAffaire )

@csrf_exempt
def suivreAffaire(request):
   if request.method == 'POST':
        try:
            data = json.loads(request.body)
            selected_ids = data.get('selected', [])
            account = request.user  # Assuming there is a one-to-one relationship with the user
            
            for id_affaire in selected_ids:
                is_suivi = SuivreAffaire.objects.filter(affaire_id=id_affaire,account=request.user)
                if not is_suivi :
                    affaire = AffaireRoles.objects.select_related('role__juridiction').get(id=id_affaire)
                    SuivreAffaire.objects.create(
                        affaire=affaire,
                        account=account,
                        juridiction=affaire.role.juridiction
                    )
            messages.success(request, 'Félicitation! Vous suivez désormais ces affaires.')
            return JsonResponse({'status': 'success'}, status=200)
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
   return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=400)


@csrf_exempt
def NePasSuivreAffaire(request):
   if request.method == 'POST':
        try:
            data = json.loads(request.body)
            selected_ids = data.get('selected', [])
            #juridiction_id = data.get('juridiction_id')
            account = request.user  # Assuming there is a one-to-one relationship with the user
            
            for id_affaire in selected_ids:
                is_suivi = SuivreAffaire.objects.filter(affaire_id=id_affaire,account=account)
                if is_suivi :
                    is_suivi.delete()
                    
            messages.success(request, 'Vous ne suivez plus ces affaires.')
            return JsonResponse({'status': 'success'}, status=200)
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
   return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=400)
   

def update_decision(request):
    if request.method == 'POST':
        decision_id = request.POST.get('decision_id')
        decision = get_object_or_404(Decisions, id=decision_id)

        # Historique modification
        old = Decisions.objects.get(id=decision.id)
        DecisionHistory.objects.create(
            original=old,
            juridiction=old.juridiction,
            affaire=old.affaire,
            decision=old.decision,
            section=old.section,
            typeDecision=old.typeDecision,
            objet=old.objet,
            president=old.president,
            greffier=old.greffier,
            dateDecision=old.dateDecision,
            dispositif=old.dispositif,
            prochaineAudience=old.prochaineAudience,
            modified_by=request.user
        )

        # Modification decision
        decision.typeDecision = request.POST.get('typeDecision')
        decision.dateDecision = request.POST.get('dateDecision') or None
        decision.decision = request.POST.get('decision')
        decision.dateDecision = request.POST.get('dateDecision')
        decision.prochaineAudience = request.POST.get('prochaineAudience') or None
        decision.statut = 'Modifier'

        decision.save()

        messages.success(request, "Décision mise à jour avec succès.")
        return redirect(request.META.get('HTTP_REFERER', '/'))

    messages.error(request, "Méthode non autorisée.")
    return redirect('/')


def ges_message(request, pk=None, action=None):
    messages = MessageDefilant.objects.all().order_by('-date_creation')

    # Créer ou éditer
    if pk and action == 'edit':
        instance = get_object_or_404(MessageDefilant, pk=pk)
        form = MessageForm(request.POST or None, instance=instance)
        form_title = "Modifier le message"
    else:
        instance = None
        form = MessageForm(request.POST or None)
        form_title = "Ajouter un message"

    if request.method == 'POST':
        if 'delete_id' in request.POST:
            message_to_delete = get_object_or_404(MessageDefilant, pk=request.POST['delete_id'])
            message_to_delete.delete()
            return redirect('ges_message')

        if form.is_valid():
            form.save()
            return redirect('ges_message')

    return render(request, 'role/gestion-message.html', {
        'messages': messages,
        'form': form,
        'form_title': form_title,
        'edit_mode': pk and action == 'edit',
        'instance_id': instance.pk if instance else None,
    })


def historique_modifications_enrollement(request, pk):
    historiques = EnrollementHistory.objects.filter(original_id=pk)  # objet_id = ID lié
    return render(request, 'role/histo_modif_enrollements.html', {'page_obj': historiques, 'original_id': pk})

def historique_modifications_decisions(request, pk):
    historiques = DecisionHistory.objects.filter(original_id=pk)  # objet_id = ID lié
    return render(request, 'role/histo_modif_decisions.html', {'historiques': historiques, 'original_id': pk})


def get_static_path(relative_path: str) -> str:
    """
    Retourne le chemin absolu d'un fichier statique pour l'utiliser (ex: ReportLab).
    relative_path : chemin relatif depuis le dossier 'static', ex: "_base/assets_role/statics/armoirie.png"
    """
    # 1. Cherche dans STATICFILES_DIRS
    for static_dir in getattr(settings, "STATICFILES_DIRS", []):
        abs_path = os.path.join(static_dir, relative_path)
        if os.path.exists(abs_path):
            return abs_path

    # 2. Cherche dans STATIC_ROOT (cas collectstatic en prod)
    abs_path = os.path.join(settings.STATIC_ROOT, relative_path)
    if os.path.exists(abs_path):
        return abs_path

    # 3. Cherche dans chaque app (utile si tu n’as pas encore fait collectstatic)
    for app in settings.INSTALLED_APPS:
        app_path = os.path.join(settings.BASE_DIR, app, "static", relative_path)
        if os.path.exists(app_path):
            return app_path

    raise FileNotFoundError(f"Static file not found: {relative_path}")


def export_roles_excel(request):
    year = request.GET.get('year')
    query = request.GET.get('q', '').strip()
    type_audience = request.GET.get('typeAudience', '').strip()
    section = request.GET.get('section', '').strip()
    get_date = request.GET.get('date', '').strip()

    roles = Roles.objects.all()
    if year:
        roles = roles.filter(dateEnreg__year=year)
    if type_audience:
        roles = roles.filter(typeAudience=type_audience)
    if section:
        roles = roles.filter(section=section)
    if get_date:
        roles = roles.filter(dateEnreg=get_date)
    if query:
        roles = roles.filter(
            Q(typeAudience__icontains=query) |
            Q(section__icontains=query) |
            Q(president__icontains=query) |
            Q(greffier__icontains=query) |
            Q(dateEnreg__icontains=query)
        )

    # Création du fichier Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Rôles"

    # En-têtes
    ws.append(["No", "Type d'audience", "Section", "Président(e)", "Greffier(e)", "Date d'enregistrement"])

    # Lignes
    for i, role in enumerate(roles, start=1):
        ws.append([
            i,
            role.typeAudience,
            role.section,
            role.president,
            role.greffier,
            role.dateEnreg.strftime('%Y-%m-%d') if role.dateEnreg else '',
        ])

    # Réponse
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=roles.xlsx'
    wb.save(response)
    return response

def export_roles_pdf(request):
    # =======================
    # Filtres
    # =======================
    year = request.GET.get('year')
    query = request.GET.get('q', '').strip()
    type_audience = request.GET.get('typeAudience', '').strip()
    section = request.GET.get('section', '').strip()
    get_date = request.GET.get('date', '').strip()

    roles = Roles.objects.all()
    if year:
        roles = roles.filter(dateEnreg__year=year)
    if type_audience:
        roles = roles.filter(typeAudience=type_audience)
    if section:
        roles = roles.filter(section=section)
    if get_date:
        roles = roles.filter(dateEnreg=get_date)
    if query:
        roles = roles.filter(
            Q(typeAudience__icontains=query) |
            Q(section__icontains=query) |
            Q(president__icontains=query) |
            Q(greffier__icontains=query) |
            Q(dateEnreg__icontains=query)
        )

    # =======================
    # Réponse HTTP
    # =======================
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="roles_{year or "tous"}.pdf"'

     # =======================
    # Callback pour pied de page
    # =======================
    def add_footer(canvas, doc):
        footer_text = f"Téléchargé à partir de Judicalex - {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
        canvas.saveState()
        canvas.setFont('Helvetica', 8)
        canvas.drawCentredString(A4[1]/2, 5 * mm, footer_text)  # centré horizontalement, 10 mm du bas
        canvas.restoreState()


    # =======================
    # Document PDF
    # =======================
    doc = SimpleDocTemplate(
        response,
        pagesize=landscape(A4),
        rightMargin=10, leftMargin=10, topMargin=20, bottomMargin=20
    )

    elements = []

    # =======================
    # Styles
    # =======================
    styles = getSampleStyleSheet()
    style_normal = styles["Normal"]
    style_normal.fontSize = 9
    style_normal.leading = 11
    style_normal.alignment = TA_CENTER

    style_title = ParagraphStyle(
        "title",
        parent=styles["Heading2"],
        alignment=TA_CENTER,  # centré
        textColor=colors.HexColor("#000000")
    )

    # =======================
    # En-tête
    # =======================
    try:
        armoirie = Image(get_static_path("_base/assets_role/statics/armoirie.png"), width=50, height=50)
        armoirie.hAlign = 'CENTER'

        branding = Image(get_static_path("_base/assets_role/statics/branding.png"), width=70, height=40)
        branding.hAlign = 'CENTER'

        simandou = Image(get_static_path("_base/assets_role/statics/simandou.png"), width=70, height=40)
        simandou.hAlign = 'CENTER'

        judicalex = Image(get_static_path("_base/assets_role/statics/ejustice_logo_white.png"), width=120, height=30)
        judicalex.hAlign = 'CENTER'

    except Exception:
        armoirie = Paragraph("[Armoirie manquante]", style_normal)
        branding = simandou = Paragraph("[Image manquante]", style_normal)
        judicalex = Paragraph("[Image manquante]", style_normal)

    # === Colonnes avec contenu centré verticalement et horizontalement ===
    col_gauche = Table(
        [[armoirie],
         [Paragraph("<b>République de Guinée</b>", style_normal)],
         [Paragraph("Travail - Justice - Solidarité", style_normal)],
         [Paragraph("Ministère de la Justice et des Droits de l'Homme", style_normal)],
         [branding]],
        style=TableStyle([
            ("ALIGN", (0,0), (-1,-1), "CENTER"),
            ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
            ("BOTTOMPADDING", (0,0), (-1,-1), 2),
            ("TOPPADDING", (0,0), (-1,-1), 2),
        ])
    )

    col_centre = Table(
        [[Paragraph("<b>COUR D'APPEL DE CONAKRY</b>", style_title)],
         [Paragraph("Tribunal de Commerce de Conakry", style_title)]],
        style=TableStyle([
            ("ALIGN", (0,0), (-1,-1), "CENTER"),
            ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
        ])
    )

    col_droite = Table(
        [[judicalex],
         [Paragraph(
             "<b>Conception & Réalisation</b><br/>"
             "Judicalex SARL<br/>"
             "contact@judicalex-gn.org<br/>"
             "Tel: 613 87 08 92", style_normal)]],
        style=TableStyle([
            ("ALIGN", (0,0), (-1,-1), "CENTER"),
            ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
        ])
    )

    # === Table principale de l'entête avec colonnes de même largeur ===
    header_table = Table(
        [[col_gauche, col_centre, col_droite]],
        colWidths=[250, 250, 250],  # même dimension pour les 3 colonnes
        rowHeights=120  # hauteur uniforme pour que tout soit centré verticalement
    )
    header_table.setStyle(TableStyle([
        ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
        ("ALIGN", (0,0), (-1,-1), "CENTER"),
        ("GRID", (0,0), (-1,-1), 0, colors.white),  # invisible, juste pour structure
    ]))

    elements.append(header_table)
    elements.append(Spacer(1, 12))

    # =======================
    # Titre du tableau avec filtres
    # =======================
    titre_filters = []

    if year:
        titre_filters.append(f"Année: {year}")
    if type_audience:
        titre_filters.append(f"Type: {type_audience}")
    if section:
        titre_filters.append(f"Section: {section}")
    if get_date:
        titre_filters.append(f"Date: {get_date}")
    if query:
        titre_filters.append(f"Recherche: {query}")

    # Construire le texte final
    titre = f"LISTE DES RÔLES"
    filtre = f"FILTRE :"
    if filtre:
        filtre += " – " + ", ".join(titre_filters)

    elements.append(Paragraph(titre, styles['Title']))
    elements.append(Paragraph(filtre, styles['Normal']))
    elements.append(Spacer(1, 12))

    # =======================
    # Tableau des rôles
    # =======================
    data = [[
        Paragraph("No", styles["Heading5"]),
        Paragraph("Type d'audience", styles["Heading5"]),
        Paragraph("Section", styles["Heading5"]),
        Paragraph("Président(e)", styles["Heading5"]),
        Paragraph("Greffier(e)", styles["Heading5"]),
        Paragraph("Date d'audience", styles["Heading5"]),
    ]]

    for i, r in enumerate(roles, start=1):
        data.append([
            Paragraph(str(i), style_normal),
            Paragraph(r.typeAudience or "", style_normal),
            Paragraph(r.section or "", style_normal),
            Paragraph(r.president or "", style_normal),
            Paragraph(r.greffier or "", style_normal),
            Paragraph(r.dateEnreg.strftime('%d/%m/%Y') if r.dateEnreg else '', style_normal),
        ])

    col_widths = [40, 150, 120, 160, 160, 120]
    table = Table(data, repeatRows=1, colWidths=col_widths)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#4CAF50')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,0), 9),
        ('BOTTOMPADDING', (0,0), (-1,0), 6),
        ('BACKGROUND', (0,1), (-1,-1), colors.whitesmoke),
        ('GRID', (0,0), (-1,-1), 0.3, colors.grey),
    ]))

    elements.append(table)

    # =======================
    # Génération du PDF
    # =======================
    doc.build(elements, onFirstPage=add_footer, onLaterPages=add_footer)

    return response


def export_roleDetail_pdf(request):
    from datetime import datetime
    # =======================
    # Récupérer les filtres
    # =======================
    query = request.GET.get('q', '').strip()
    role_id = request.GET.get('role_id', '').strip()

    # Base queryset
    affaire = AffaireRoles.objects.filter(role_id=role_id)
    role = Roles.objects.filter(id=role_id).first()

    if query:
        affaire = affaire.filter(
            Q(objet__icontains=query) |
            Q(demandeurs__icontains=query) |
            Q(defendeurs__icontains=query)
        )

    # =======================
    # Réponse HTTP PDF
    # =======================
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="role_{role.dateEnreg or "tous"}.pdf"'

    # =======================
    # Callback pour pied de page
    # =======================
    def add_footer(canvas, doc):
        footer_text = f"Téléchargé à partir de Judicalex - {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
        canvas.saveState()
        canvas.setFont('Helvetica', 8)
        canvas.drawCentredString(A4[1]/2, 5 * mm, footer_text)
        canvas.restoreState()

    # =======================
    # Document PDF
    # =======================
    doc = SimpleDocTemplate(
        response,
        pagesize=landscape(A4),
        rightMargin=10, leftMargin=10, topMargin=20, bottomMargin=20
    )

    elements = []

    # =======================
    # Styles
    # =======================
    styles = getSampleStyleSheet()
    style_normal = styles["Normal"]
    style_normal.fontSize = 9
    style_normal.leading = 11
    style_normal.alignment = TA_CENTER

    style_title = ParagraphStyle(
        "title",
        parent=styles["Heading2"],
        alignment=TA_CENTER,
        textColor=colors.HexColor("#000000")
    )

    # =======================
    # En-tête (logos et textes)
    # =======================
    try:
        armoirie = Image(get_static_path("_base/assets_role/statics/armoirie.png"), width=50, height=50)
        armoirie.hAlign = 'CENTER'

        branding = Image(get_static_path("_base/assets_role/statics/branding.png"), width=70, height=40)
        branding.hAlign = 'CENTER'

        simandou = Image(get_static_path("_base/assets_role/statics/simandou.png"), width=70, height=40)
        simandou.hAlign = 'CENTER'

        judicalex = Image(get_static_path("_base/assets_role/statics/ejustice_logo_white.png"), width=120, height=30)
        judicalex.hAlign = 'CENTER'

    except Exception:
        armoirie = Paragraph("[Armoirie manquante]", style_normal)
        branding = simandou = Paragraph("[Image manquante]", style_normal)
        judicalex = Paragraph("[Image manquante]", style_normal)

    col_gauche = Table(
        [[armoirie],
         [Paragraph("<b>République de Guinée</b>", style_normal)],
         [Paragraph("Travail - Justice - Solidarité", style_normal)],
         [Paragraph("Ministère de la Justice et des Droits de l'Homme", style_normal)],
         [branding]],
        style=TableStyle([
            ("ALIGN", (0,0), (-1,-1), "CENTER"),
            ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
            ("BOTTOMPADDING", (0,0), (-1,-1), 2),
            ("TOPPADDING", (0,0), (-1,-1), 2),
        ])
    )

    col_centre = Table(
        [[Paragraph("<b>COUR D'APPEL DE CONAKRY</b>", style_title)],
         [Paragraph("Tribunal de Commerce de Conakry", style_title)]],
        style=TableStyle([
            ("ALIGN", (0,0), (-1,-1), "CENTER"),
            ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
        ])
    )

    col_droite = Table(
        [[judicalex],
         [Paragraph(
             "<b>Conception & Réalisation</b><br/>"
             "Judicalex SARL<br/>"
             "contact@judicalex-gn.org<br/>"
             "Tel: 613 87 08 92", style_normal)]],
        style=TableStyle([
            ("ALIGN", (0,0), (-1,-1), "CENTER"),
            ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
        ])
    )

    header_table = Table(
        [[col_gauche, col_centre, col_droite]],
        colWidths=[250, 250, 250],
        rowHeights=120
    )
    header_table.setStyle(TableStyle([
        ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
        ("ALIGN", (0,0), (-1,-1), "CENTER"),
        ("GRID", (0,0), (-1,-1), 0, colors.white),
    ]))
    elements.append(header_table)
    elements.append(Spacer(1, 12))

    # =======================
    # Titre avec filtres
    # =======================
    titre_filters = []
    if query:
        titre_filters.append(f"Recherche: {query}")

    titre_final = f"RÔLE D'AUDIENCE DU {role.typeAudience} DU {role.dateEnreg.strftime('%d/%m/%Y') if role.dateEnreg else 'Tous'}"
    filtre_text = ""
    if titre_filters:
        filtre_text = "FILTRE – " + ", ".join(titre_filters)

    elements.append(Paragraph(titre_final, styles['Title']))
    if filtre_text:
        elements.append(Paragraph(filtre_text, style_normal))
    elements.append(Spacer(1, 12))

    # =======================
    # Tableau des affaires
    # =======================
    data = [[
        Paragraph("No", styles["Heading5"]),
        Paragraph("NUA", styles["Heading5"]),
        Paragraph("RG", styles["Heading5"]),
        Paragraph("Demandeurs", styles["Heading5"]),
        Paragraph("Défendeurs", styles["Heading5"]),
        Paragraph("Objet", styles["Heading5"]),
    ]]

    for i, r in enumerate(affaire, start=1):
        data.append([
            Paragraph(str(i), style_normal),
            Paragraph(r.numAffaire or "", style_normal),
            Paragraph(r.numRg or "", style_normal),
            Paragraph(r.demandeurs or "", style_normal),
            Paragraph(r.defendeurs or "", style_normal),
            Paragraph(r.objet or "", style_normal),
        ])

    col_widths = [40, 150, 120, 160, 160, 120]
    table = Table(data, repeatRows=1, colWidths=col_widths)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#4CAF50')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,0), 9),
        ('BOTTOMPADDING', (0,0), (-1,0), 6),
        ('BACKGROUND', (0,1), (-1,-1), colors.whitesmoke),
        ('GRID', (0,0), (-1,-1), 0.3, colors.grey),
    ]))

    elements.append(table)

    # =======================
    # Génération du PDF
    # =======================
    doc.build(elements, onFirstPage=add_footer, onLaterPages=add_footer)

    return response

def export_roleDetail_excel(request):
    query = request.GET.get('q', '').strip()
    role_id = request.GET.get('role_id', '').strip()

    affaire = AffaireRoles.objects.filter(role_id=role_id)
    audience = Roles.objects.filter(id=role_id).first()
   
    if query:
        affaire = affaire.filter(
            Q(typeAudience__icontains=query) |
            Q(section__icontains=query) |
            Q(president__icontains=query) |
            Q(greffier__icontains=query) |
            Q(dateEnreg__icontains=query)
        )

    # Création du fichier Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Rôles"

    # En-têtes
    ws.append(["No", "NUA", "RG", "Demandeurs", "Defendeurs", "Objet"])

    # Lignes
    for i, role in enumerate(affaire, start=1):
        ws.append([
            i,
            role.numAffaire,
            role.numRg,
            role.demandeurs,
            role.defendeurs,
            role.objet,
        ])

    # Réponse
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=roles_{audience.dateEnreg}.xlsx'
    wb.save(response)
    return response


def export_decisions_pdf(request):
    from datetime import datetime

    # =======================
    # Récupérer les filtres
    # =======================
    affaire_id = request.GET.get('affaire_id', '').strip()
    affaire = AffaireRoles.objects.filter(id=affaire_id).first()

    if not affaire:
        return HttpResponse("Affaire introuvable", status=404)

    decisions = Decisions.objects.select_related('affaire').filter(
        affaire__objet=affaire.objet,
        affaire__demandeurs=affaire.demandeurs,
        affaire__defendeurs=affaire.defendeurs,
    )

    # =======================
    # Réponse HTTP PDF
    # =======================
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="decisions_affaire_{affaire.numAffaire or "tous"}.pdf"'

    # =======================
    # Callback pour pied de page
    # =======================
    def add_footer(canvas, doc):
        footer_text = f"Téléchargé à partir de Judicalex - {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
        canvas.saveState()
        canvas.setFont('Helvetica', 8)
        canvas.drawCentredString(A4[1]/2, 5 * mm, footer_text)
        canvas.restoreState()

    # =======================
    # Document PDF
    # =======================
    doc = SimpleDocTemplate(
        response,
        pagesize=landscape(A4),
        rightMargin=10, leftMargin=10, topMargin=20, bottomMargin=20
    )

    elements = []

    # =======================
    # Styles
    # =======================
    styles = getSampleStyleSheet()
    style_normal = styles["Normal"]
    style_normal.fontSize = 9
    style_normal.leading = 11
    style_normal.alignment = TA_CENTER

    style_title = ParagraphStyle(
        "title",
        parent=styles["Heading2"],
        alignment=TA_CENTER,
        textColor=colors.HexColor("#000000")
    )

    # =======================
    # En-tête (logos et textes)
    # =======================
    try:
        armoirie = Image(get_static_path("_base/assets_role/statics/armoirie.png"), width=50, height=50)
        armoirie.hAlign = 'CENTER'

        branding = Image(get_static_path("_base/assets_role/statics/branding.png"), width=70, height=40)
        branding.hAlign = 'CENTER'

        simandou = Image(get_static_path("_base/assets_role/statics/simandou.png"), width=70, height=40)
        simandou.hAlign = 'CENTER'

        judicalex = Image(get_static_path("_base/assets_role/statics/ejustice_logo_white.png"), width=120, height=30)
        judicalex.hAlign = 'CENTER'

    except Exception:
        armoirie = Paragraph("[Armoirie manquante]", style_normal)
        branding = simandou = Paragraph("[Image manquante]", style_normal)
        judicalex = Paragraph("[Image manquante]", style_normal)

    col_gauche = Table(
        [[armoirie],
         [Paragraph("<b>République de Guinée</b>", style_normal)],
         [Paragraph("Travail - Justice - Solidarité", style_normal)],
         [Paragraph("Ministère de la Justice et des Droits de l'Homme", style_normal)],
         [branding]],
        style=TableStyle([
            ("ALIGN", (0,0), (-1,-1), "CENTER"),
            ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
            ("BOTTOMPADDING", (0,0), (-1,-1), 2),
            ("TOPPADDING", (0,0), (-1,-1), 2),
        ])
    )

    col_centre = Table(
        [[Paragraph("<b>COUR D'APPEL DE CONAKRY</b>", style_title)],
         [Paragraph("Tribunal de Commerce de Conakry", style_title)]],
        style=TableStyle([
            ("ALIGN", (0,0), (-1,-1), "CENTER"),
            ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
        ])
    )

    col_droite = Table(
        [[judicalex],
         [Paragraph(
             "<b>Conception & Réalisation</b><br/>"
             "Judicalex SARL<br/>"
             "contact@judicalex-gn.org<br/>"
             "Tel: 613 87 08 92", style_normal)]],
        style=TableStyle([
            ("ALIGN", (0,0), (-1,-1), "CENTER"),
            ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
        ])
    )

    header_table = Table(
        [[col_gauche, col_centre, col_droite]],
        colWidths=[250, 250, 250],
        rowHeights=120
    )
    header_table.setStyle(TableStyle([
        ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
        ("ALIGN", (0,0), (-1,-1), "CENTER"),
        ("GRID", (0,0), (-1,-1), 0, colors.white),
    ]))
    elements.append(header_table)
    elements.append(Spacer(1, 12))

    # =======================
    # Titre avec filtre
    # =======================
    titre = f"DÉCISIONS DE L'AFFAIRE"
    sous_titre = f"AFFAIRE N° {affaire.numAffaire} - {affaire.objet}"
    elements.append(Paragraph(titre, styles['Title']))
    elements.append(Paragraph(sous_titre, styles['Normal']))
    elements.append(Spacer(1, 12))

    # =======================
    # Tableau des décisions
    # =======================
    data = [[
        Paragraph("No", styles["Heading5"]),
        Paragraph("Audience du", styles["Heading5"]),
        Paragraph("Type de décision", styles["Heading5"]),
        Paragraph("Décision", styles["Heading5"]),
        Paragraph("Prochaine Audience", styles["Heading5"]),
    ]]

    for i, r in enumerate(decisions, start=1):
        data.append([
            Paragraph(str(i), style_normal),
            Paragraph(r.dateDecision.strftime("%d/%m/%Y") if r.dateDecision else "", style_normal),
            Paragraph(r.typeDecision or "", style_normal),
            Paragraph(r.decision or "", style_normal),
            Paragraph(r.prochaineAudience.strftime("%d/%m/%Y") if r.prochaineAudience else "", style_normal),
        ])

    col_widths = [40, 150, 120, 160, 160]
    table = Table(data, repeatRows=1, colWidths=col_widths)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#4CAF50')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,0), 9),
        ('BOTTOMPADDING', (0,0), (-1,0), 6),
        ('BACKGROUND', (0,1), (-1,-1), colors.whitesmoke),
        ('GRID', (0,0), (-1,-1), 0.3, colors.grey),
    ]))

    elements.append(table)

    # =======================
    # Génération du PDF
    # =======================
    doc.build(elements, onFirstPage=add_footer, onLaterPages=add_footer)

    return response

def export_decisions_excel(request):
    # Récupérer les filtres
    affaire_id = request.GET.get('affaire_id', '').strip()

    # Base queryset
    affaire = AffaireRoles.objects.filter(id=affaire_id).first()

    decisions = Decisions.objects.select_related('affaire').filter(
        affaire__objet=affaire.objet,
        affaire__demandeurs=affaire.demandeurs,
        affaire__defendeurs=affaire.defendeurs,
    )

    # Création du fichier Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Rôles"

    # En-têtes
    ws.append(["No", "Audience du", "Type de decision", "Decision", "Prochaine audience"])

    # Lignes
    for i, aff in enumerate(decisions, start=1):
        ws.append([
            i,
            aff.dateDecision.strftime("%d/%m/%Y"),
            aff.typeDecision,
            aff.decision,
            aff.prochaineAudience.strftime("%d/%m/%Y"),
        ])

    # Réponse
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=decisions_affaire_{affaire.numAffaire}.xlsx'
    wb.save(response)
    return response


def export_plumitifs_excel(request):

    year = request.GET.get('year')
    query = request.GET.get('q', '').strip()
    type_audience = request.GET.get('typeAudience', '').strip()
    section = request.GET.get('section', '').strip()
    get_date = request.GET.get('date', '').strip()

    affaireRole = AffaireRoles.objects.all()
    if year:
        affaireRole = affaireRole.filter(role__dateEnreg__year=year)
    if type_audience:
        affaireRole = affaireRole.filter(role__typeAudience=type_audience)
    if section:
        affaireRole = affaireRole.filter(role__section=section)
    if get_date:
        affaireRole = affaireRole.filter(role__dateEnreg=get_date)

    if query:
        affaireRole = affaireRole.filter(
            Q(numRg__icontains=query) |
            Q(demandeurs__icontains=query) |
            Q(defendeurs__icontains=query) |
            Q(objet__icontains=query)
        )

    # Création Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PLUMITIF"

    ws.append(["No", "NUA", "RG", "Demanderesse", "Défenderesse", "Objet", "date d'audience", "Décisions"])
    for i, affaire in enumerate(affaireRole, start=1):
        decision_text = Decisions.objects.filter(affaire=affaire).first() or ""
        ws.append([
            i,
            affaire.numAffaire,
            affaire.numRg,
            affaire.demandeurs,
            affaire.defendeurs,
            affaire.objet,
            affaire.role.dateEnreg,
            decision_text.decision,
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=affaires.xlsx'
    wb.save(response)
    return response

def export_plumitifs_pdf(request):
    from datetime import datetime

    # =======================
    # Filtres
    # =======================
    year = request.GET.get('year')
    query = request.GET.get('q', '').strip()
    type_audience = request.GET.get('typeAudience', '').strip()
    section = request.GET.get('section', '').strip()
    get_date = request.GET.get('date', '').strip()

    affaireRole = AffaireRoles.objects.all()
    if year:
        affaireRole = affaireRole.filter(role__dateEnreg__year=year)
    if type_audience:
        affaireRole = affaireRole.filter(role__typeAudience=type_audience)
    if section:
        affaireRole = affaireRole.filter(role__section=section)
    if get_date:
        affaireRole = affaireRole.filter(role__dateEnreg=get_date)

    if query:
        affaireRole = affaireRole.filter(
            Q(numRg__icontains=query) |
            Q(demandeurs__icontains=query) |
            Q(defendeurs__icontains=query) |
            Q(objet__icontains=query)
        )

    # =======================
    # Réponse HTTP PDF
    # =======================
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="affaires_{year or "tous"}.pdf"'

    # =======================
    # Callback pour pied de page
    # =======================
    def add_footer(canvas, doc):
        footer_text = f"Téléchargé à partir de Judicalex - {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
        canvas.saveState()
        canvas.setFont('Helvetica', 8)
        canvas.drawCentredString(A4[1]/2, 5 * mm, footer_text)
        canvas.restoreState()

    # =======================
    # Document PDF
    # =======================
    doc = SimpleDocTemplate(response, pagesize=landscape(A4),
                            rightMargin=10, leftMargin=10, topMargin=20, bottomMargin=20)
    elements = []
    styles = getSampleStyleSheet()
    style_normal = styles["Normal"]
    style_normal.fontSize = 9
    style_normal.leading = 11
    style_normal.alignment = TA_CENTER

    style_title = ParagraphStyle(
        "title",
        parent=styles["Heading2"],
        alignment=TA_CENTER,
        textColor=colors.HexColor("#000000")
    )

    # =======================
    # En-tête (logos et textes)
    # =======================
    try:
        armoirie = Image(get_static_path("_base/assets_role/statics/armoirie.png"), width=50, height=50)
        armoirie.hAlign = 'CENTER'

        branding = Image(get_static_path("_base/assets_role/statics/branding.png"), width=70, height=40)
        branding.hAlign = 'CENTER'

        simandou = Image(get_static_path("_base/assets_role/statics/simandou.png"), width=70, height=40)
        simandou.hAlign = 'CENTER'

        judicalex = Image(get_static_path("_base/assets_role/statics/ejustice_logo_white.png"), width=120, height=30)
        judicalex.hAlign = 'CENTER'

    except Exception:
        armoirie = Paragraph("[Armoirie manquante]", style_normal)
        branding = simandou = Paragraph("[Image manquante]", style_normal)
        judicalex = Paragraph("[Image manquante]", style_normal)

    col_gauche = Table(
        [[armoirie],
         [Paragraph("<b>République de Guinée</b>", style_normal)],
         [Paragraph("Travail - Justice - Solidarité", style_normal)],
         [Paragraph("Ministère de la Justice et des Droits de l'Homme", style_normal)],
         [branding]],
        style=TableStyle([
            ("ALIGN", (0,0), (-1,-1), "CENTER"),
            ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
            ("BOTTOMPADDING", (0,0), (-1,-1), 2),
            ("TOPPADDING", (0,0), (-1,-1), 2),
        ])
    )

    col_centre = Table(
        [[Paragraph("<b>COUR D'APPEL DE CONAKRY</b>", style_title)],
         [Paragraph("Tribunal de Commerce de Conakry", style_title)]],
        style=TableStyle([
            ("ALIGN", (0,0), (-1,-1), "CENTER"),
            ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
        ])
    )

    col_droite = Table(
        [[judicalex],
         [Paragraph(
             "<b>Conception & Réalisation</b><br/>"
             "Judicalex SARL<br/>"
             "contact@judicalex-gn.org<br/>"
             "Tel: 613 87 08 92", style_normal)]],
        style=TableStyle([
            ("ALIGN", (0,0), (-1,-1), "CENTER"),
            ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
        ])
    )

    header_table = Table(
        [[col_gauche, col_centre, col_droite]],
        colWidths=[250, 250, 250],
        rowHeights=120
    )
    header_table.setStyle(TableStyle([
        ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
        ("ALIGN", (0,0), (-1,-1), "CENTER"),
        ("GRID", (0,0), (-1,-1), 0, colors.white),
    ]))
    elements.append(header_table)
    elements.append(Spacer(1, 12))

    # =======================
    # Titre avec filtres
    # =======================
    titre_filters = []
    if year:
        titre_filters.append(f"Année: {year}")
    if type_audience:
        titre_filters.append(f"Type: {type_audience}")
    if section:
        titre_filters.append(f"Section: {section}")
    if get_date:
        titre_filters.append(f"Date: {get_date}")
    if query:
        titre_filters.append(f"Recherche: {query}")

    titre = f"PLUMITIF"
    filtre = ""
    if titre_filters:
        filtre = "FILTRE : " + ", ".join(titre_filters)

    elements.append(Paragraph(titre, styles['Title']))
    if filtre:
        elements.append(Paragraph(filtre, styles['Normal']))
    elements.append(Spacer(1, 12))

    # =======================
    # Tableau des affaires
    # =======================
    data = [[
        Paragraph("No", styles["Heading5"]),
        Paragraph("NUA", styles["Heading5"]),
        Paragraph("RG", styles["Heading5"]),
        Paragraph("Demanderesse", styles["Heading5"]),
        Paragraph("Défenderesse", styles["Heading5"]),
        Paragraph("Objet", styles["Heading5"]),
        Paragraph("Date d'audience", styles["Heading5"]),
        Paragraph("Décisions", styles["Heading5"])
    ]]

    for i, affaire in enumerate(affaireRole, start=1):

        decision_text = Decisions.objects.filter(affaire=affaire).first() or ""

        data.append([
            Paragraph(str(i), style_normal),
            Paragraph(affaire.numAffaire or "", style_normal),
            Paragraph(affaire.numRg or "", style_normal),
            Paragraph(affaire.demandeurs or "", style_normal),
            Paragraph(affaire.defendeurs or "", style_normal),
            Paragraph(affaire.objet or "", style_normal),
            Paragraph(affaire.role.dateEnreg.strftime("%d/%m/%Y") if affaire.role and affaire.role.dateEnreg else "", style_normal),
            Paragraph(decision_text.decision, style_normal),
        ])

    col_widths = [30, 80, 60, 120, 120, 90, 80, 110]
    table = Table(data, repeatRows=1, colWidths=col_widths)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4CAF50')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
        ('BACKGROUND', (0, 1), (-1, -1), colors.whitesmoke),
        ('GRID', (0, 0), (-1, -1), 0.3, colors.grey),
    ]))

    elements.append(table)

    # =======================
    # Génération PDF
    # =======================
    doc.build(elements, onFirstPage=add_footer, onLaterPages=add_footer)

    return response


def export_enrollements_excel(request):

    # Récupérer les filtres
    year = request.GET.get('year')
    query = request.GET.get('q', '').strip()
    type_audience = request.GET.get('typeAudience', '').strip()
    get_date = request.GET.get('date', '').strip()

    # Base queryset
    enrollements = Enrollement.objects.all()
    if year:
        enrollements = enrollements.filter(dateEnrollement__year=year)
    if type_audience:
        enrollements = enrollements.filter(typeAudience=type_audience)
    if get_date:
        enrollements = enrollements.filter(dateEnrollement=get_date)
    if query:
        enrollements = enrollements.filter(
            Q(typeAudience__icontains=query) |
            Q(demandeurs__icontains=query) |
            Q(defendeurs__icontains=query) |
            Q(objet__icontains=query) |
            Q(numRg__icontains=query) |
            Q(numAffaire__icontains=query)
    )

    # Créer un fichier Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Enrôlements"

    # En-têtes
    ws.append([
        "No", "RG", "Type d'audience", "Section",
        "Date d'enrôlement", "Date d'audience",
        "Demanderesse", "Défenderesse", "Objet"
    ])

    # Contenu
    for i, e in enumerate(enrollements, start=1):
        ws.append([
            i,
            e.numRg,
            e.typeAudience,
            e.section,
            e.dateEnrollement.strftime('%Y-%m-%d') if e.dateEnrollement else '',
            e.dateAudience.strftime('%Y-%m-%d') if e.dateAudience else '',
            e.demandeurs,
            e.defendeurs,
            e.objet
        ])

    # Réponse HTTP avec fichier Excel
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"enrollements_{year}.xlsx"
    response['Content-Disposition'] = f'attachment; filename={filename}'

    wb.save(response)
    return response

def export_enrollements_pdf(request):

    # =======================
    # Filtres
    # =======================
    year = request.GET.get('year')
    query = request.GET.get('q', '').strip()
    type_audience = request.GET.get('typeAudience', '').strip()
    get_date = request.GET.get('date', '').strip()

    # Base queryset
    enrollements = Enrollement.objects.all()
    if year:
        enrollements = enrollements.filter(dateEnrollement__year=year)
    if type_audience:
        enrollements = enrollements.filter(typeAudience=type_audience)
    if get_date:
        enrollements = enrollements.filter(dateEnrollement=get_date)
    if query:
        enrollements = enrollements.filter(
            Q(typeAudience__icontains=query) |
            Q(demandeurs__icontains=query) |
            Q(defendeurs__icontains=query) |
            Q(objet__icontains=query) |
            Q(numRg__icontains=query) |
            Q(numAffaire__icontains=query)
        )

    # =======================
    # Réponse HTTP PDF
    # =======================
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="enrollements_{year or "tous"}.pdf"'

    # =======================
    # Callback pour pied de page
    # =======================
    def add_footer(canvas, doc):
        footer_text = f"Téléchargé à partir de Judicalex - {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
        canvas.saveState()
        canvas.setFont('Helvetica', 8)
        canvas.drawCentredString(A4[1]/2, 5 * mm, footer_text)
        canvas.restoreState()

    # =======================
    # Document PDF
    # =======================
    doc = SimpleDocTemplate(response, pagesize=landscape(A4),
                            rightMargin=10, leftMargin=10, topMargin=20, bottomMargin=20)
    elements = []
    styles = getSampleStyleSheet()
    style_normal = styles["Normal"]
    style_normal.fontSize = 9
    style_normal.leading = 11
    style_normal.alignment = TA_CENTER

    style_title = ParagraphStyle(
        "title",
        parent=styles["Heading2"],
        alignment=TA_CENTER,
        textColor=colors.HexColor("#000000")
    )

    # =======================
    # En-tête (logos et textes)
    # =======================
    try:
        armoirie = Image(get_static_path("_base/assets_role/statics/armoirie.png"), width=50, height=50)
        armoirie.hAlign = 'CENTER'

        branding = Image(get_static_path("_base/assets_role/statics/branding.png"), width=70, height=40)
        branding.hAlign = 'CENTER'

        simandou = Image(get_static_path("_base/assets_role/statics/simandou.png"), width=70, height=40)
        simandou.hAlign = 'CENTER'

        judicalex = Image(get_static_path("_base/assets_role/statics/ejustice_logo_white.png"), width=120, height=30)
        judicalex.hAlign = 'CENTER'

    except Exception:
        armoirie = Paragraph("[Armoirie manquante]", style_normal)
        branding = simandou = Paragraph("[Image manquante]", style_normal)
        judicalex = Paragraph("[Image manquante]", style_normal)

    col_gauche = Table(
        [[armoirie],
         [Paragraph("<b>République de Guinée</b>", style_normal)],
         [Paragraph("Travail - Justice - Solidarité", style_normal)],
         [Paragraph("Ministère de la Justice et des Droits de l'Homme", style_normal)],
         [branding]],
        style=TableStyle([
            ("ALIGN", (0,0), (-1,-1), "CENTER"),
            ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
            ("BOTTOMPADDING", (0,0), (-1,-1), 2),
            ("TOPPADDING", (0,0), (-1,-1), 2),
        ])
    )

    col_centre = Table(
        [[Paragraph("<b>COUR D'APPEL DE CONAKRY</b>", style_title)],
         [Paragraph("Tribunal de Commerce de Conakry", style_title)]],
        style=TableStyle([
            ("ALIGN", (0,0), (-1,-1), "CENTER"),
            ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
        ])
    )

    col_droite = Table(
        [[judicalex],
         [Paragraph(
             "<b>Conception & Réalisation</b><br/>"
             "Judicalex SARL<br/>"
             "contact@judicalex-gn.org<br/>"
             "Tel: 613 87 08 92", style_normal)]],
        style=TableStyle([
            ("ALIGN", (0,0), (-1,-1), "CENTER"),
            ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
        ])
    )

    header_table = Table(
        [[col_gauche, col_centre, col_droite]],
        colWidths=[250, 250, 250],
        rowHeights=120
    )
    header_table.setStyle(TableStyle([
        ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
        ("ALIGN", (0,0), (-1,-1), "CENTER"),
        ("GRID", (0,0), (-1,-1), 0, colors.white),
    ]))
    elements.append(header_table)
    elements.append(Spacer(1, 12))

    # =======================
    # Titre avec filtres
    # =======================
    titre_filters = []
    if year:
        titre_filters.append(f"Année: {year}")
    if type_audience:
        titre_filters.append(f"Type: {type_audience}")
    if get_date:
        titre_filters.append(f"Date: {get_date}")
    if query:
        titre_filters.append(f"Recherche: {query}")

    titre = f"REGISTRE D'ENROLLEMENTS"
    filtre = ""
    if titre_filters:
        filtre = "FILTRE : " + ", ".join(titre_filters)

    elements.append(Paragraph(titre, styles['Title']))
    if filtre:
        elements.append(Paragraph(filtre, styles['Normal']))
    elements.append(Spacer(1, 12))

    # =======================
    # Tableau des enrollements
    # =======================
    data = [
        [
            Paragraph("NUA", styles["Heading5"]),
            Paragraph("RG", styles["Heading5"]),
            Paragraph("Date Enrôlement", styles["Heading5"]),
            Paragraph("Date Audience", styles["Heading5"]),
            Paragraph("Demanderesse", styles["Heading5"]),
            Paragraph("Défenderesse", styles["Heading5"]),
            Paragraph("Objet", styles["Heading5"])
        ]
    ]

    for e in enrollements:
        data.append([
            Paragraph(e.numAffaire or "", style_normal),
            Paragraph(e.numRg or "", style_normal),
            Paragraph(e.dateEnrollement.strftime('%d/%m/%Y') if e.dateEnrollement else '', style_normal),
            Paragraph(e.dateAudience.strftime('%d/%m/%Y') if e.dateAudience else '', style_normal),
            Paragraph(e.demandeurs or "", style_normal),
            Paragraph(e.defendeurs or "", style_normal),
            Paragraph(e.objet or "", style_normal)
        ])

    col_widths = [60, 50, 80, 80, 130, 130, 250]
    table = Table(data, repeatRows=1, colWidths=col_widths)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#4CAF50')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,0), 9),
        ('BOTTOMPADDING', (0,0), (-1,0), 6),
        ('BACKGROUND', (0,1), (-1,-1), colors.whitesmoke),
        ('GRID', (0,0), (-1,-1), 0.3, colors.grey),
    ]))

    elements.append(table)

    # =======================
    # Génération PDF
    # =======================
    doc.build(elements, onFirstPage=add_footer, onLaterPages=add_footer)

    return response

def check_doublon(request):
    numRg = request.GET.get('numRg', '').strip()
    demandeurs = request.GET.get('demandeurs', '').strip()
    defendeurs = request.GET.get('defendeurs', '').strip()
    objet = request.GET.get('objet', '').strip()

    # On cherche des enregistrements ressemblants
    doublons = Enrollement.objects.filter(
        (Q(demandeurs__icontains=demandeurs) & Q(defendeurs__icontains=defendeurs))
    ).values('id', 'numRg', 'demandeurs', 'defendeurs', 'objet')[:5]

    return JsonResponse({
        'has_doublon': doublons.exists(),
        'doublons': list(doublons)
    })